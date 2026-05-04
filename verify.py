#!/usr/bin/env python3
"""
verify.py — HTML ↔ XLSX sync check for elvinci_auftragssteuerung
Run after any data change to JTL_DATA or JTL_Rechnungstexte.xlsx.

Usage: python3 verify.py [path-to-html]   # default: ./elvinci_auftragssteuerung.html
Requirements: pip install openpyxl
"""
import re, json, sys, subprocess
from pathlib import Path

HTML_PATH = Path(sys.argv[1] if len(sys.argv) > 1 else 'elvinci_auftragssteuerung.html')
XLSX_PATH = Path('data/JTL_Rechnungstexte.xlsx')

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("✗ openpyxl missing. Run: pip install openpyxl")

if not HTML_PATH.exists(): sys.exit(f"✗ HTML not found: {HTML_PATH}")
if not XLSX_PATH.exists(): sys.exit(f"✗ XLSX not found: {XLSX_PATH}")

html = HTML_PATH.read_text(encoding='utf-8')

# ─── 1. JTL_DATA parses cleanly ─────────────────────────────────
m = re.search(r'const JTL_DATA = (\[.*?\n\]);\s*\n', html, re.DOTALL)
if not m: sys.exit("✗ JTL_DATA constant not found in HTML")
data = json.loads(m.group(1))
print(f"✓ JTL_DATA parses — {len(data)} entries")

# ─── 2. Shape invariants ─────────────────────────────────────────
assert len(data) == 45, f"Expected 45 entries, got {len(data)}"
assert all('counter' not in e for e in data), "counter field present (eliminated in v3 — bug)"
print("✓ Shape: 45 entries, no counter dimension")

# ─── 3. JS syntax check (if node available) ─────────────────────
try:
    js_match = re.search(r'<script>\n(.*?)\n</script>', html, re.DOTALL)
    if js_match:
        tmp = Path('/tmp/_elvinci_check.js')
        tmp.write_text(js_match.group(1))
        r = subprocess.run(['node', '--check', str(tmp)], capture_output=True, text=True)
        if r.returncode == 0:
            print("✓ JS syntax valid (node --check)")
        else:
            sys.exit(f"✗ JS syntax error:\n{r.stderr}")
except FileNotFoundError:
    print("⚠ node not found — skipping JS syntax check")

# ─── 4. HTML JSON ↔ XLSX field-level diff ───────────────────────
wb = load_workbook(XLSX_PATH, data_only=True)
sheet_meta = {
    'Lieferung OHNE Terminwunsch': ('Lieferung','ohne'),
    'Lieferung MIT Terminwunsch':  ('Lieferung','mit'),
    'Abholungen OHNE Terminwunsch':('Abholung','ohne'),
    'Abholungen MIT Terminwunsch': ('Abholung','mit'),
}
def map_status(s):
    if s is None: return None
    s = str(s).strip().upper()
    if 'THEORETISCH' in s and 'NICHT' in s: return 'theoretisch_nicht_erfuellbar'
    if 'THEORETISCH' in s: return 'theoretisch_erfuellbar'
    if 'NICHT' in s: return 'nicht_erfuellbar'
    if 'ERFÜLLBAR' in s: return 'erfuellbar'
    return None

xlsx_data = []
for sn, (va, tw) in sheet_meta.items():
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    has_termin = (tw == 'mit'); has_hebe = (va == 'Lieferung')
    start = next(i for i, r in enumerate(rows, start=1)
                 if r and any(str(c or '').strip().upper() == 'OPTIONEN' for c in r))
    for r in rows[start:]:
        if not any(v is not None and str(v).strip() for v in r): continue
        col = 2
        tw_status = map_status(r[col]) if has_termin else None
        if has_termin: col += 1
        hebe = None
        if has_hebe:
            hebe = ('mit' if r[col] and 'MIT' in str(r[col]).upper()
                    else 'ohne' if r[col] and 'OHNE' in str(r[col]).upper() else None)
            col += 1
        za = str(r[col]).strip().upper() if r[col] else None; col += 1
        zm = re.search(r'(\d+)', str(r[col]) if r[col] else '')
        zz = zm.group(1) if zm else None; col += 1
        rest = [v for v in r[col:] if v is not None]
        xlsx_data.append({
            'versandart': va, 'terminwunsch': tw, 'terminwunsch_status': tw_status,
            'hebebuehne': hebe, 'zahlungsart': za, 'zahlungsziel': zz,
            'todos_vor':       str(rest[0]).rstrip() if len(rest) > 0 else '',
            'jtl_text':        str(rest[1]).rstrip() if len(rest) > 1 else '',
            'todos_nach':      str(rest[2]).rstrip() if len(rest) > 2 else '',
            'dashboard_text':  str(rest[3]).rstrip() if len(rest) > 3 else '',
        })

assert len(xlsx_data) == 45, f"XLSX has {len(xlsx_data)} entries, expected 45"

def keyof(d):
    return (d['versandart'], d['terminwunsch'], d['terminwunsch_status'],
            d['hebebuehne'], d['zahlungsart'], d['zahlungsziel'])

html_by_key = {keyof(d): d for d in data}
xlsx_by_key = {keyof(d): d for d in xlsx_data}
missing_in_html = set(xlsx_by_key) - set(html_by_key)
missing_in_xlsx = set(html_by_key) - set(xlsx_by_key)
assert not missing_in_html, f"Keys missing in HTML: {missing_in_html}"
assert not missing_in_xlsx, f"Keys missing in XLSX: {missing_in_xlsx}"

diffs = 0
for k in html_by_key:
    h, x = html_by_key[k], xlsx_by_key[k]
    for f in ['todos_vor', 'jtl_text', 'todos_nach', 'dashboard_text']:
        ht = re.sub(r' {2,}', ' ', (h.get(f) or '').strip())
        xt = re.sub(r' {2,}', ' ', (x.get(f) or '').strip())
        if ht != xt:
            diffs += 1
            if diffs <= 3:
                print(f"\n  DIFF in {k} / {f}:")
                print(f"    HTML: {ht[:120]!r}")
                print(f"    XLSX: {xt[:120]!r}")

assert diffs == 0, f"\n✗ {diffs} text fields differ between HTML and XLSX"
print(f"✓ HTML JSON ↔ XLSX — 0 diffs across {len(html_by_key) * 4} fields")

# ─── 5. Logic test: every entry findable via simulated matching ─
def simulate(params):
    heberStr = 'mit' if params['hebebuehne'] else 'ohne'
    zzStr = str(params['zahlungsziel']) if params['zahlungsziel'] is not None else None
    for s in data:
        if s['versandart'] != params['versandart']: continue
        if s['terminwunsch'] != params['terminwunsch']: continue
        if params['terminwunsch'] == 'mit' and s['terminwunsch_status'] != params['terminwunsch_status']: continue
        if params['versandart'] == 'Lieferung' and s['hebebuehne'] != heberStr: continue
        if s['zahlungsart'] != params['zahlungsart']: continue
        if params['zahlungsart'] == 'ZAHLUNGSZIEL' and str(s['zahlungsziel']) != zzStr: continue
        return s
    return None

found = 0
for entry in data:
    p = {
        'versandart': entry['versandart'],
        'terminwunsch': entry['terminwunsch'],
        'terminwunsch_status': entry['terminwunsch_status'],
        'hebebuehne': entry['hebebuehne'] == 'mit' if entry['hebebuehne'] else False,
        'zahlungsart': entry['zahlungsart'],
        'zahlungsziel': entry['zahlungsziel'],
    }
    res = simulate(p)
    if res and keyof(res) == keyof(entry):
        found += 1

assert found == 45, f"Only {found}/45 entries findable via simulated matching"
print(f"✓ All 45 entries match uniquely via findJTLText simulation")

print("\n" + "=" * 50)
print("ALL CHECKS PASSED")
print("=" * 50)
